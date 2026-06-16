"""
logger.py
─────────────────────────────────────────────
Thread-safe event logger with in-memory ring buffer, CSV and Excel export.
"""

import csv
import logging
import sqlite3
import threading
from collections import deque
from datetime import datetime
from pathlib import Path
from typing import Callable, Deque, Dict, List, Optional

import config

# ── stdlib logger ──────────────────────────────────────────────────────────
logging.basicConfig(
    level=getattr(logging, config.LOG_LEVEL, logging.INFO),
    format="%(asctime)s  %(levelname)-8s  %(name)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
_log = logging.getLogger("DDS")


class EventRecord:
    """Single log entry."""

    __slots__ = ("timestamp", "event_type", "driver", "ear", "detail")

    def __init__(
        self,
        event_type: str,
        driver: str = "",
        ear: float = 0.0,
        detail: str = "",
    ) -> None:
        self.timestamp  = datetime.now()
        self.event_type = event_type
        self.driver     = driver
        self.ear        = ear
        self.detail     = detail

    # ── helpers ─────────────────────────────────────────────────────────
    def ts_str(self) -> str:
        return self.timestamp.strftime("%Y-%m-%d %H:%M:%S")

    def to_row(self) -> List[str]:
        return [self.ts_str(), self.event_type, self.driver,
                f"{self.ear:.4f}", self.detail]

    def to_display(self) -> str:
        return f"[{self.ts_str()}]  {self.event_type:<20}  EAR={self.ear:.3f}  {self.detail}"


class EventLogger:
    """
    Central logger for the drowsiness detection system.

    Features
    ────────
    • Thread-safe ring buffer (deque) for UI display
    • Persistent SQLite store for history/statistics
    • CSV / Excel export
    • Observer callbacks for real-time UI updates
    """

    _CSV_HEADER = ["Timestamp", "EventType", "Driver", "EAR", "Detail"]

    def __init__(self) -> None:
        self._lock: threading.Lock = threading.Lock()
        self._buffer: Deque[EventRecord] = deque(maxlen=config.MAX_LOG_ENTRIES)
        self._callbacks: List[Callable[[EventRecord], None]] = []
        self._db: Optional[sqlite3.Connection] = None
        self._init_csv()
        self._init_db()

    # ── init ────────────────────────────────────────────────────────────
    def _init_csv(self) -> None:
        path = config.LOG_CSV_FILE
        if not path.exists():
            with open(path, "w", newline="") as f:
                csv.writer(f).writerow(self._CSV_HEADER)

    def _init_db(self) -> None:
        try:
            self._db = sqlite3.connect(
                str(config.DRIVER_DB_FILE), check_same_thread=False
            )
            self._db.execute(
                """CREATE TABLE IF NOT EXISTS events (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp  TEXT,
                    event_type TEXT,
                    driver     TEXT,
                    ear        REAL,
                    detail     TEXT
                )"""
            )
            self._db.commit()
        except sqlite3.Error as exc:
            _log.error("DB init failed: %s", exc)

    # ── public API ──────────────────────────────────────────────────────
    def register_callback(self, fn: Callable[[EventRecord], None]) -> None:
        """Register a UI callback that fires on every new event."""
        with self._lock:
            self._callbacks.append(fn)

    def log(
        self,
        event_type: str,
        driver: str = "",
        ear: float = 0.0,
        detail: str = "",
    ) -> None:
        record = EventRecord(event_type, driver, ear, detail)
        _log.info("%s | %s | %s | %.3f | %s", record.ts_str(),
                  event_type, driver, ear, detail)
        with self._lock:
            self._buffer.append(record)
            callbacks = list(self._callbacks)
        self._persist(record)
        for fn in callbacks:
            try:
                fn(record)
            except Exception as exc:               # noqa: BLE001
                _log.debug("Callback error: %s", exc)

    def get_recent(self, n: int = 100) -> List[EventRecord]:
        with self._lock:
            return list(self._buffer)[-n:]

    def get_all_db(self) -> List[Dict]:
        if not self._db:
            return []
        try:
            cur = self._db.execute(
                "SELECT timestamp,event_type,driver,ear,detail FROM events"
                " ORDER BY id DESC LIMIT 1000"
            )
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, row)) for row in cur.fetchall()]
        except sqlite3.Error:
            return []

    # ── persistence ─────────────────────────────────────────────────────
    def _persist(self, rec: EventRecord) -> None:
        # CSV (append)
        try:
            with open(config.LOG_CSV_FILE, "a", newline="") as f:
                csv.writer(f).writerow(rec.to_row())
        except OSError as exc:
            _log.debug("CSV write error: %s", exc)
        # SQLite
        if self._db:
            try:
                self._db.execute(
                    "INSERT INTO events(timestamp,event_type,driver,ear,detail)"
                    " VALUES(?,?,?,?,?)",
                    (rec.ts_str(), rec.event_type, rec.driver,
                     rec.ear, rec.detail),
                )
                self._db.commit()
            except sqlite3.Error as exc:
                _log.debug("DB write error: %s", exc)

    # ── export ──────────────────────────────────────────────────────────
    def export_csv(self, path: Optional[Path] = None) -> Path:
        """Export full log buffer to a CSV file and return the path."""
        out = path or config.LOG_CSV_FILE
        with self._lock:
            rows = list(self._buffer)
        with open(out, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(self._CSV_HEADER)
            for rec in rows:
                w.writerow(rec.to_row())
        _log.info("Exported %d records → %s", len(rows), out)
        return out

    def export_excel(self, path: Optional[Path] = None) -> Path:
        """Export full log buffer to an Excel file and return the path."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            _log.warning("openpyxl not installed — falling back to CSV export")
            return self.export_csv(path)

        out = path or config.LOG_EXCEL_FILE
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "DrowsinessLog"

        # Header styling
        hdr_fill = PatternFill("solid", fgColor="1C1C1C")
        hdr_font = Font(bold=True, color="00D4FF")
        for col, h in enumerate(self._CSV_HEADER, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        with self._lock:
            rows = list(self._buffer)
        for r_idx, rec in enumerate(rows, 2):
            for c_idx, val in enumerate(rec.to_row(), 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(out)
        _log.info("Excel export → %s", out)
        return out

    def close(self) -> None:
        if self._db:
            try:
                self._db.close()
            except sqlite3.Error:
                pass


# ── module-level singleton ─────────────────────────────────────────────────
event_logger = EventLogger()
